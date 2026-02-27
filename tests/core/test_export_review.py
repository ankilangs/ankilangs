"""Tests for export_review — CSV and Excel generation for native speaker review."""

import csv
from pathlib import Path

import pytest

from al_tools.core import csv2sqlite, export_review

# Shared testdata with 15 entries (enough to test empty-row-every-10 logic)
_TESTDATA_DIR = Path(__file__).parent / "testdata" / "test_export_review"


@pytest.fixture()
def review_db(tmp_path):
    """Create a test DB from shared export_review testdata."""
    db_path = tmp_path / "test.db"
    csv2sqlite(_TESTDATA_DIR, db_path, force=True)
    return db_path


def test_export_review_csv_structure(review_db, tmp_path):
    """CSV has correct headers, row count, and content."""
    export_review(
        review_db,
        source_locale="en_us",
        target_locale="es_es",
        output_dir=tmp_path / "output",
        media_dir=tmp_path / "nonexistent_audio",
        data_dir=_TESTDATA_DIR,
    )

    csv_file = tmp_path / "output" / "review_en_us_to_es_es.csv"
    assert csv_file.exists()

    with open(csv_file, encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        rows = list(reader)

    # Check headers
    expected_headers = [
        "guid",
        "clarification",
        "source_text",
        "target_text",
        "target_ipa",
        "pronunciation_hint",
        "spelling_hint",
        "reading_hint",
        "listening_hint",
        "notes",
        "review_comment",
    ]
    assert reader.fieldnames == expected_headers

    # Non-empty rows should contain our 15 data entries
    data_rows = [r for r in rows if r["guid"]]
    assert len(data_rows) == 15

    # Spot-check first row (alphabetically: "the apple")
    first = data_rows[0]
    assert first["source_text"] == "the apple"
    assert first["target_text"] == "la manzana"
    assert first["target_ipa"] == "/la manˈθana/"
    assert first["review_comment"] == ""


def test_export_review_csv_empty_rows(review_db, tmp_path):
    """Empty rows inserted every 10 entries (at positions 10, 20, ...)."""
    export_review(
        review_db,
        source_locale="en_us",
        target_locale="es_es",
        output_dir=tmp_path / "output",
        media_dir=tmp_path / "nonexistent_audio",
        data_dir=_TESTDATA_DIR,
    )

    csv_file = tmp_path / "output" / "review_en_us_to_es_es.csv"
    with open(csv_file, encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        rows = list(reader)

    # With 15 data entries, there should be 1 empty row (after entry 10)
    # Total rows = 15 data + 1 empty = 16
    empty_rows = [r for r in rows if not r["guid"]]
    assert len(empty_rows) == 1
    assert len(rows) == 16


def test_export_review_excel_created(review_db, tmp_path):
    """Excel file is created with correct structure."""
    export_review(
        review_db,
        source_locale="en_us",
        target_locale="es_es",
        output_dir=tmp_path / "output",
        media_dir=tmp_path / "nonexistent_audio",
        data_dir=_TESTDATA_DIR,
    )

    xlsx_file = tmp_path / "output" / "review_en_us_to_es_es.xlsx"
    assert xlsx_file.exists()

    from openpyxl import load_workbook

    wb = load_workbook(xlsx_file, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # First row is headers
    headers = rows[0]
    assert "guid" in headers
    assert "review_comment" in headers
    assert "target_text" in headers


def test_export_review_no_pairs(review_db, tmp_path, capsys):
    """Prints message and returns when no translation pairs found."""
    export_review(
        review_db,
        source_locale="en_us",
        target_locale="xx_xx",  # nonexistent pair
        output_dir=tmp_path / "output",
        media_dir=tmp_path / "audio",
        data_dir=_TESTDATA_DIR,
    )

    captured = capsys.readouterr()
    assert "No translation pairs found" in captured.out

    # No files should be created
    output_dir = tmp_path / "output"
    csv_files = list(output_dir.glob("*.csv")) if output_dir.exists() else []
    assert len(csv_files) == 0


def test_export_review_no_audio_dir(review_db, tmp_path, capsys):
    """Skips audio concatenation when audio dir missing, still produces CSV+Excel."""
    export_review(
        review_db,
        source_locale="en_us",
        target_locale="es_es",
        output_dir=tmp_path / "output",
        media_dir=tmp_path / "nonexistent_audio",
        data_dir=_TESTDATA_DIR,
    )

    captured = capsys.readouterr()
    assert "Audio directory" in captured.out
    assert "not found" in captured.out

    # CSV and Excel should still exist
    assert (tmp_path / "output" / "review_en_us_to_es_es.csv").exists()
    assert (tmp_path / "output" / "review_en_us_to_es_es.xlsx").exists()

    # No audio file should exist
    mp3_files = list((tmp_path / "output").glob("*.mp3"))
    assert len(mp3_files) == 0


def test_export_review_with_keys_filter(review_db, tmp_path):
    """Only entries matching the given keys are exported."""
    selected_keys = ["the cat", "the dog", "the fish"]
    export_review(
        review_db,
        source_locale="en_us",
        target_locale="es_es",
        output_dir=tmp_path / "output",
        media_dir=tmp_path / "nonexistent_audio",
        data_dir=_TESTDATA_DIR,
        keys=selected_keys,
    )

    csv_file = tmp_path / "output" / "review_en_us_to_es_es.csv"
    with open(csv_file, encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        rows = list(reader)

    data_rows = [r for r in rows if r["guid"]]
    assert len(data_rows) == 3
    exported_sources = {r["source_text"] for r in data_rows}
    assert exported_sources == {"the cat", "the dog", "the fish"}


def test_export_review_with_keys_filter_no_match(review_db, tmp_path, capsys):
    """Keys that don't exist produce 'No translation pairs found' message."""
    export_review(
        review_db,
        source_locale="en_us",
        target_locale="es_es",
        output_dir=tmp_path / "output",
        media_dir=tmp_path / "nonexistent_audio",
        data_dir=_TESTDATA_DIR,
        keys=["nonexistent_key"],
    )

    captured = capsys.readouterr()
    assert "No translation pairs found" in captured.out


def test_export_review_with_empty_keys_list(review_db, tmp_path, capsys):
    """Empty keys list exports nothing."""
    export_review(
        review_db,
        source_locale="en_us",
        target_locale="es_es",
        output_dir=tmp_path / "output",
        media_dir=tmp_path / "nonexistent_audio",
        data_dir=_TESTDATA_DIR,
        keys=[],
    )

    captured = capsys.readouterr()
    assert "No translation pairs found" in captured.out
